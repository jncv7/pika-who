import boto3
import json
import openpyxl
from openpyxl import Workbook
import time
import datetime
from datetime import datetime
from colorama import Fore, Back, Style

# Initialize AWS IAM client
iam = boto3.client('iam')

def get_usernames():
    """Retrieve all IAM usernames with pagination support."""
    usernames = []
    response = iam.list_users()
    while True:
        usernames.extend([user['UserName'] for user in response['Users']])
        if response.get('IsTruncated'):
            # There are more users to retrieve
            response = iam.list_users(Marker=response['Marker'])
        else:
            break
    return usernames

def get_managed_policies(username):
    """Retrieve managed policies attached to a user with pagination support."""
    policies = []
    response = iam.list_attached_user_policies(UserName=username)
    while True:
        policies.extend([(policy['PolicyArn'], policy['PolicyName']) for policy in response['AttachedPolicies']])
        if response.get('IsTruncated'):
            # There are more policies to retrieve
            response = iam.list_attached_user_policies(UserName=username, Marker=response['Marker'])
        else:
            break
    return policies

def get_inline_policies(username):
    """Retrieve inline policies attached to a user with pagination support."""
    policies = []
    response = iam.list_user_policies(UserName=username)
    while True:
        policies.extend(response['PolicyNames'])
        if response.get('IsTruncated'):
            # There are more policies to retrieve
            response = iam.list_user_policies(UserName=username, Marker=response['Marker'])
        else:
            break
    return policies

def get_managed_policy_details(policy_arn):
    """Retrieve details of a managed policy."""
    policy = iam.get_policy(PolicyArn=policy_arn)
    version_id = policy['Policy']['DefaultVersionId']
    
    # Get the policy version details
    version = iam.get_policy_version(PolicyArn=policy_arn, VersionId=version_id)
    return json.dumps(version['PolicyVersion']['Document'], separators=(',', ':'))

def get_inline_policy_details(username, policy_name):
    """Retrieve details of an inline policy attached to a user."""
    response = iam.get_user_policy(UserName=username, PolicyName=policy_name)
    return json.dumps(response['PolicyDocument'], separators=(',', ':'))

def filter_s3_iam_permissions(policy_document):
    """Check policy document for S3 and IAM permissions and return relevant actions if found."""
    s3_actions = []
    iam_actions = []
    
    # Parse the policy document JSON
    policy_statements = json.loads(policy_document).get("Statement", [])
    if not isinstance(policy_statements, list):
        policy_statements = [policy_statements]  # Ensure it's a list for consistency
    
    for statement in policy_statements:
        actions = statement.get("Action", [])
        if isinstance(actions, str):
            actions = [actions]  # Ensure actions is a list

        for action in actions:
            if action.startswith("s3:"):
                s3_actions.append(action)
            elif action.startswith("iam:"):
                iam_actions.append(action)

    return s3_actions, iam_actions

def write_to_excel(users_data):
    """Write the IAM user policies data to an Excel file."""
    # Initialize workbook and sheets
    workbook = Workbook()
    managed_sheet = workbook.active
    managed_sheet.title = "Managed Policies"
    inline_sheet = workbook.create_sheet(title="Inline Policies")
    summary_sheet = workbook.create_sheet(title="User Summary")
    s3_iam_sheet = workbook.create_sheet(title="S3 and IAM Permissions")
    
    # Headers for managed, inline, summary, and S3/IAM permission sheets
    # This also effects the order of the tabs.
    summary_sheet.append(["UserName", "Managed Policy Names", "Inline Policy Names"])
    managed_sheet.append(["UserName", "Managed Policy Name", "Policy ARN", "Policy Details"])
    inline_sheet.append(["UserName", "Inline Policy Name", "Policy Details"])
    s3_iam_sheet.append(["Policy Name", "Policy ARN", "S3 Permissions", "IAM Permissions"])
    
    # Write data for managed policies
    for user in users_data:
        username = user['UserName']
        managed_policy_names = []
        inline_policy_names = []
        
        # Managed policies
        for policy in user['ManagedPolicies']:
            managed_sheet.append([
                username,
                policy['PolicyName'],
                policy['PolicyArn'],
                policy['PolicyDetails']
            ])
            managed_policy_names.append(policy['PolicyName'])
            
            # Check for S3 and IAM permissions
            s3_actions, iam_actions = filter_s3_iam_permissions(policy['PolicyDetails'])
            if s3_actions or iam_actions:
                s3_iam_sheet.append([
                    policy['PolicyName'],
                    policy['PolicyArn'],
                    ", ".join(s3_actions) if s3_actions else "None",
                    ", ".join(iam_actions) if iam_actions else "None"
                ])

        # Inline policies
        for policy in user['InlinePolicies']:
            inline_sheet.append([
                username,
                policy['PolicyName'],
                policy['PolicyDetails']
            ])
            inline_policy_names.append(policy['PolicyName'])

            # Check for S3 and IAM permissions
            s3_actions, iam_actions = filter_s3_iam_permissions(policy['PolicyDetails'])
            if s3_actions or iam_actions:
                s3_iam_sheet.append([
                    policy['PolicyName'],
                    "Inline Policy",
                    ", ".join(s3_actions) if s3_actions else "None",
                    ", ".join(iam_actions) if iam_actions else "None"
                ])

        # Write summary information
        summary_sheet.append([
            username,
            ", ".join(managed_policy_names) if managed_policy_names else "None",
            ", ".join(inline_policy_names) if inline_policy_names else "None"
        ])

    # Save the workbook
    filename = f"IAM_Users_Policies_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    workbook.save(filename)
    print(f"Data written to {filename}")

def main():

    print (Fore.GREEN + " --- Start Time ---")
    print (datetime.now())
    print(Style.RESET_ALL)
    

    print("Starting IAM user policies collection...")
    print ("")
    

    # List to hold all users and their policy data
    users_data = []
    time.sleep (3)


    # Retrieve all usernames
    usernames = get_usernames()
    for username in usernames:
        print(f"Processing user: {username}")
        user_data = {
            'UserName': username,
            'ManagedPolicies': [],
            'InlinePolicies': []
        }
        
        # Retrieve managed policies
        managed_policies = get_managed_policies(username)
        for policy_arn, policy_name in managed_policies:
            policy_details = get_managed_policy_details(policy_arn)
            user_data['ManagedPolicies'].append({
                'PolicyName': policy_name,
                'PolicyArn': policy_arn,
                'PolicyDetails': policy_details
            })

        # Retrieve inline policies
        inline_policies = get_inline_policies(username)
        for policy_name in inline_policies:
            policy_details = get_inline_policy_details(username, policy_name)
            user_data['InlinePolicies'].append({
                'PolicyName': policy_name,
                'PolicyDetails': policy_details
            })

        users_data.append(user_data)

    # Write the collected data to Excel
    write_to_excel(users_data)
    print ("Reporting is Complete. Please check Host Folder for the Export File.")

    print (Fore.GREEN + ' --- End Time --- ')
    print (datetime.now())
    print(Style.RESET_ALL)
    print ("")
    print ("")
    print (Fore.RED + "Please take a screenshot of this window for IPE. This window will close shortly ...!")
    print(Style.RESET_ALL)
    #delays the window from closing so user can take a screenshot for IPE
    time.sleep (10)

if __name__ == '__main__':
    main()